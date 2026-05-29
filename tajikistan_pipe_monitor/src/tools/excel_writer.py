import os
import json
import logging
from datetime import datetime
from typing import List, Dict, Any, Union, Optional
from crewai.tools import BaseTool
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

logger = logging.getLogger(__name__)

class ExcelWriterTool(BaseTool):
    name: str = "Excel Registry Writer"
    description: str = (
        "Создаёт Excel-реестр проектов, текстовое резюме и файл контактов в папке output/. "
        "Параметры: "
        "projects_json (JSON-строка со списком проектов), "
        "summary_text (необязательный текст резюме на русском), "
        "contacts_json (необязательный JSON-строка с контактами ПИУ/ПМУ)."
    )

    def _run(self, projects_json: str, summary_text: Optional[str] = None, contacts_json: Optional[str] = None) -> str:
        try:
            # Create output dir if not exist
            output_dir = "output"
            os.makedirs(output_dir, exist_ok=True)
            date_str = datetime.now().strftime("%Y%m%d")

            # 1. Write Excel file
            excel_path = self._write_excel(projects_json, output_dir, date_str)
            result_msg = f"Excel реестр успешно сохранен: {excel_path}"

            # 2. Write summary text if provided
            if summary_text:
                summary_path = os.path.join(output_dir, f"urgent_summary_{date_str}.txt")
                latest_summary_path = os.path.join(output_dir, "urgent_summary_latest.txt")
                with open(summary_path, "w", encoding="utf-8") as f:
                    f.write(summary_text)
                with open(latest_summary_path, "w", encoding="utf-8") as f:
                    f.write(summary_text)
                result_msg += f"\nРезюме сохранено: {summary_path}"

            # 3. Write contacts JSON if provided
            if contacts_json:
                contacts_path = os.path.join(output_dir, f"contacts_{date_str}.json")
                latest_contacts_path = os.path.join(output_dir, "contacts_latest.json")
                
                # Try parsing and formatting JSON nicely
                try:
                    parsed_contacts = json.loads(contacts_json)
                    formatted_contacts = json.dumps(parsed_contacts, ensure_ascii=False, indent=2)
                except Exception:
                    formatted_contacts = contacts_json
                    
                with open(contacts_path, "w", encoding="utf-8") as f:
                    f.write(formatted_contacts)
                with open(latest_contacts_path, "w", encoding="utf-8") as f:
                    f.write(formatted_contacts)
                result_msg += f"\nКонтакты сохранены: {contacts_path}"

            return result_msg

        except Exception as e:
            logger.error(f"Error in ExcelWriterTool: {str(e)}")
            return f"Ошибка записи отчетов: {str(e)}"

    def _write_excel(self, projects_json: str, output_dir: str, date_str: str) -> str:
        if isinstance(projects_json, str):
            try:
                data = json.loads(projects_json)
            except Exception as e:
                clean_json = projects_json.strip()
                if clean_json.startswith("```json"):
                    clean_json = clean_json[7:]
                if clean_json.endswith("```"):
                    clean_json = clean_json[:-3]
                data = json.loads(clean_json.strip())
        else:
            data = projects_json

        projects = []
        if isinstance(data, list):
            projects = data
        elif isinstance(data, dict):
            for val in data.values():
                if isinstance(val, list):
                    projects = val
                    break
            else:
                projects = [data]

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Реестр проектов"
        ws.views.sheetView[0].showGridLines = True

        col_widths = {
            "A": 6, "B": 25, "C": 25, "D": 28, "E": 28, "F": 20, "G": 18,
            "H": 12, "I": 30, "J": 30, "K": 35, "L": 38, "M": 18, "N": 15,
            "O": 12, "P": 15, "Q": 15, "R": 15, "S": 20, "T": 45
        }

        for col, width in col_widths.items():
            ws.column_dimensions[col].width = width

        font_title = Font(name="Arial", size=14, bold=True, color="FFFFFF")
        font_subtitle = Font(name="Arial", size=11, italic=True, color="000000")
        font_header_group = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        font_header = Font(name="Arial", size=10, bold=True, color="000000")
        font_data = Font(name="Arial", size=9)
        font_footer = Font(name="Arial", size=8, italic=True, color="595959")

        fill_title = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        fill_subtitle = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        fill_header_group_main = PatternFill(start_color="418AB3", end_color="418AB3", fill_type="solid")
        fill_header_group_extra = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        fill_header = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

        align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        align_left = Alignment(horizontal="left", vertical="top", wrap_text=True)
        align_center_top = Alignment(horizontal="center", vertical="top", wrap_text=True)

        border_side = Side(border_style="thin", color="D3D3D3")
        thin_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)

        # Row 1: Title
        ws.merge_cells("A1:T1")
        ws["A1"] = "РЕЕСТР ПРОЕКТОВ ПО СТРОИТЕЛЬСТВУ И РЕКОНСТРУКЦИИ ТРУБОПРОВОДОВ В ТАДЖИКИСТАНЕ"
        ws["A1"].font = font_title
        ws["A1"].fill = fill_title
        ws["A1"].alignment = align_center
        ws.row_dimensions[1].height = 40

        # Row 2: Subtitle
        ws.merge_cells("A2:T2")
        ws["A2"] = "ТЕКУЩИЕ ПРОЕКТЫ НА СТАДИИ РЕАЛИЗАЦИИ И ПОДГОТОВКИ (ДИАМЕТР DN ≥ 400 мм)"
        ws["A2"].font = font_subtitle
        ws["A2"].fill = fill_subtitle
        ws["A2"].alignment = align_center
        ws.row_dimensions[2].height = 25

        # Row 3: Super headers
        ws.merge_cells("A3:L3")
        ws["A3"] = "ОСНОВНЫЕ ПАРАМЕТРЫ ПРОЕКТОВ (ФОРМАТ COMPOSITE.TJ)"
        ws["A3"].font = font_header_group
        ws["A3"].fill = fill_header_group_main
        ws["A3"].alignment = align_center
        
        ws.merge_cells("M3:T3")
        ws["M3"] = "ТЕХНИЧЕСКИЕ И ТЕНДЕРНЫЕ ДЕТАЛИ"
        ws["M3"].font = font_header_group
        ws["M3"].fill = fill_header_group_extra
        ws["M3"].alignment = align_center
        ws.row_dimensions[3].height = 22

        # Row 4: Column Headers
        headers = [
            ("A", "№"), ("B", "Название проекта"), ("C", ""), ("D", "Исполнительное агентство"),
            ("E", "Финансирующее учреждение"), ("F", "Статус проекта"), ("G", "Срок реализации проекта"),
            ("H", "Продолжительность (мес.)"), ("I", "Компании-исполнители"), ("J", "Бюджет (доля Барса) USD"),
            ("K", "Проблемы / Риски"), ("L", "Ответственный / Контакты"), ("M", "Диаметр труб (DN, мм)"),
            ("N", "Материал труб"), ("O", "Длина (км)"), ("P", "Метод закупки"), ("Q", "Дата IFB/GPN"),
            ("R", "Дедлайн тендера"), ("S", "Пригодность для Барса"), ("T", "Ссылка на тендер")
        ]

        ws.merge_cells("B4:C4")

        for col_letter, header_text in headers:
            cell = ws[f"{col_letter}4"]
            cell.value = header_text
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center
            cell.border = thin_border
        ws.row_dimensions[4].height = 30

        for col in range(1, 21):
            ws.cell(row=3, column=col).border = thin_border
            ws.cell(row=4, column=col).border = thin_border

        status_colors = {
            "Активен": "D4EDDA", "В исполнении": "D4EDDA", "Не объявлен": "D0E8FF",
            "Ожидает": "FFF3CD", "Подготовка": "FFE4CC", "Новый — GPN": "FFE4CC",
        }

        suitability_colors = {
            "HIGH": "F8D7DA", "MEDIUM": "FFF3CD", "LOW": "FFFFFF", "NO": "E2E3E5"
        }

        row_idx = 5
        for idx, p in enumerate(projects):
            p_dict = p if isinstance(p, dict) else p.model_dump()
            ws.row_dimensions[row_idx].height = 90

            contact_parts = []
            if p_dict.get("piu_pmu"):
                contact_parts.append(f"ПМУ: {p_dict.get('piu_pmu')}")
            if p_dict.get("contact_name"):
                contact_parts.append(f"ФИО: {p_dict.get('contact_name')}")
            if p_dict.get("contact_org") and p_dict.get("contact_org") != p_dict.get("piu_pmu"):
                contact_parts.append(f"Орг: {p_dict.get('contact_org')}")
            if p_dict.get("contact_email"):
                contact_parts.append(f"Email: {p_dict.get('contact_email')}")
            if p_dict.get("contact_phone"):
                contact_parts.append(f"Тел: {p_dict.get('contact_phone')}")
            if p_dict.get("contact_address"):
                contact_parts.append(f"Адр: {p_dict.get('contact_address')}")
            
            contacts_str = "\n".join(contact_parts) if contact_parts else "нет данных"

            row_data = {
                "A": p_dict.get("num", idx + 1),
                "B": p_dict.get("name", "нет данных"),
                "C": "",
                "D": p_dict.get("executing_agency", "нет данных"),
                "E": p_dict.get("donor", "нет данных"),
                "F": p_dict.get("status", "нет данных"),
                "G": p_dict.get("period") or p_dict.get("deadline_date") or "нет данных",
                "H": p_dict.get("duration_months") or "",
                "I": p_dict.get("contractor") or "Барс (потенциально)",
                "J": p_dict.get("composite_share") or p_dict.get("total_budget") or "нет данных",
                "K": p_dict.get("risks") or "нет рисков",
                "L": contacts_str,
                "M": p_dict.get("pipe_diameter_mm") or "нет данных",
                "N": p_dict.get("pipe_material") or "нет данных",
                "O": p_dict.get("pipe_length_km") or "",
                "P": p_dict.get("procurement_method") or "нет данных",
                "Q": p_dict.get("ifb_date") or "нет данных",
                "R": p_dict.get("deadline_date") or "нет данных",
                "S": p_dict.get("suitability") or p_dict.get("urgency") or "LOW",
                "T": p_dict.get("source_url") or "нет ссылки",
            }

            ws.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=3)

            for col_letter, val in row_data.items():
                cell = ws[f"{col_letter}{row_idx}"]
                cell.value = val
                cell.font = font_data
                cell.border = thin_border
                
                if col_letter in ["A", "F", "G", "H", "M", "N", "O", "P", "Q", "R", "S"]:
                    cell.alignment = align_center_top
                else:
                    cell.alignment = align_left

            status_val = str(row_data["F"]).strip()
            status_fill_color = "FFFFFF"
            for k, color in status_colors.items():
                if k.lower() in status_val.lower():
                    status_fill_color = color
                    break
            if status_fill_color != "FFFFFF":
                ws[f"F{row_idx}"].fill = PatternFill(start_color=status_fill_color, end_color=status_fill_color, fill_type="solid")

            suit_val = str(row_data["S"]).strip().upper()
            suit_fill_color = suitability_colors.get(suit_val, "FFFFFF")
            if suit_fill_color != "FFFFFF":
                ws[f"S{row_idx}"].fill = PatternFill(start_color=suit_fill_color, end_color=suit_fill_color, fill_type="solid")

            row_idx += 1

        row_idx += 1
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=20)
        footer_val = f"Реестр сгенерирован автоматически системой CrewAI. Дата выгрузки: {datetime.now().strftime('%Y-%m-%d %H:%M')}. Источников сканировано: 10."
        ws.cell(row=row_idx, column=1).value = footer_val
        ws.cell(row=row_idx, column=1).font = font_footer
        ws.cell(row=row_idx, column=1).alignment = align_center
        ws.row_dimensions[row_idx].height = 20

        ws.freeze_panes = "A5"

        dated_filename = f"tajikistan_pipe_registry_{date_str}.xlsx"
        dated_path = os.path.join(output_dir, dated_filename)
        latest_path = os.path.join(output_dir, "tajikistan_pipe_registry_latest.xlsx")

        wb.save(dated_path)
        wb.save(latest_path)
        return dated_path
